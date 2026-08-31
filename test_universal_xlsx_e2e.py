from io import BytesIO
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from universal_integration import converter_previa_para_resultado
from validator import validar
from xlsx_writer import gerar_xlsx


def _template_minimo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Regular"
    for nome in ("Grupo de itens adicionais", "Item Pesado", "Pizza"):
        wb.create_sheet(nome)
    for sheet in wb.worksheets:
        sheet.cell(1, 1).value = "linha preservada 1"
        sheet.cell(2, 1).value = "cabecalho preservado"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _previa_aprovada():
    return SimpleNamespace(
        fonte="teste-e2e",
        validacao={"aprovado": True, "erros": [], "avisos": []},
        avisos=[],
        produtos=[
            {
                "codigo": "p1",
                "nome": "Hamburguer Teste",
                "descricao": "Produto regular",
                "categoria": "Lanches",
                "imagem": "https://exemplo.com/foto.jpg",
                "preco": 25.9,
                "grupos": ["g1"],
            },
            {
                "codigo": "pz1",
                "nome": "Pizza Teste",
                "descricao": "Pizza",
                "categoria": "Pizzas",
                "imagem": "",
                "preco": 40.0,
                "grupos": ["g1"],
                "pizza": True,
                "metodo_preco_pizza": 1,
            },
        ],
        pizzas=[{"codigo": "pz1", "pizza": True, "metodo_preco_pizza": 1}],
        grupos=[
            {
                "grupo_id": "g1",
                "tipo": 1,
                "grupo_nome": "Adicionais",
                "nome": "Bacon",
                "preco": 5.0,
                "minimo": 0,
                "maximo": 2,
                "repetir": 0,
                "metodo_preco": 1,
            }
        ],
    )


def test_v2_aprovada_chega_ao_xlsx_atual_sem_perder_vinculos():
    resultado = converter_previa_para_resultado(_previa_aprovada())
    erros, _avisos = validar(resultado)
    assert erros == []

    xlsx = gerar_xlsx(_template_minimo(), resultado)
    wb = load_workbook(BytesIO(xlsx), data_only=False)

    wi = wb["Item Regular"]
    wg = wb["Grupo de itens adicionais"]
    wp = wb["Pizza"]

    assert wi.cell(1, 1).value == "linha preservada 1"
    assert wi.cell(2, 1).value == "cabecalho preservado"
    assert wi.cell(3, 3).value == "Hamburguer Teste"
    assert wi.cell(3, 2).value == "10001"
    assert wi.cell(3, 8).value == 25.9

    assert wg.cell(3, 1).value == 10001
    assert wg.cell(3, 3).value == "Adicionais"
    assert wg.cell(3, 4).value == "Bacon"
    assert wg.cell(3, 6).value == 5.0

    assert wp.cell(3, 3).value == "Pizza Teste"
    assert wp.cell(3, 2).value == "10001"
    assert wp.cell(3, 7).value == 1
    assert wp.cell(3, 8).value == 40.0


def test_xlsx_gerado_nao_cria_formulas_ou_hyperlinks():
    resultado = converter_previa_para_resultado(_previa_aprovada())
    xlsx = gerar_xlsx(_template_minimo(), resultado)
    wb = load_workbook(BytesIO(xlsx), data_only=False)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.hyperlink is None
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("=")


def test_pizza_sem_metodo_de_preco_bloqueia_xlsx_no_validador():
    previa = _previa_aprovada()
    previa.produtos[1]["metodo_preco_pizza"] = 0
    previa.pizzas[0]["metodo_preco_pizza"] = 0

    resultado = converter_previa_para_resultado(previa)
    erros, _avisos = validar(resultado)

    assert any("sem método de preço definido" in erro for erro in erros)
