from io import BytesIO
from copy import copy
from urllib.parse import urlsplit
from openpyxl import load_workbook

ITEM_HEADERS = 21
GRUPO_HEADERS = 10
PIZZA_HEADERS = 20


def _copy_row_style(ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d._style = copy(s._style)
        if s.number_format:
            d.number_format = s.number_format
        if s.font:
            d.font = copy(s.font)
        if s.fill:
            d.fill = copy(s.fill)
        if s.border:
            d.border = copy(s.border)
        if s.alignment:
            d.alignment = copy(s.alignment)
        if s.protection:
            d.protection = copy(s.protection)


def _clear_data(ws, max_col):
    if ws.max_row >= 3:
        for row in ws.iter_rows(
            min_row=3,
            max_row=ws.max_row,
            min_col=1,
            max_col=max_col,
        ):
            for cell in row:
                cell.value = None
                if cell.hyperlink:
                    cell.hyperlink = None


def _imagem_para_importacao(url):
    """
    O importador aceita melhor URLs que aparentam possuir extensão de imagem.
    Mantemos a URL original e acrescentamos apenas um parâmetro inofensivo
    quando ela não termina em jpg/jpeg/png.
    """
    if not url:
        return ""

    u = str(url).strip()
    if not u:
        return ""

    try:
        path = (urlsplit(u).path or "").lower()
    except Exception:
        path = u.lower().split("?")[0].split("#")[0]

    if path.endswith((".jpg", ".jpeg", ".png")):
        return u

    if path.endswith(".webp"):
        # Não fingimos que WEBP é JPG.
        return ""

    sep = "&" if "?" in u else "?"
    return f"{u}{sep}img=.jpg"


def _montar_mapa_grupos(resultado):
    """
    O template/importador trabalha com códigos NUMÉRICOS.
    As plataformas modernas usam UUID/ObjectId/string.
    Convertemos os IDs externos para inteiros locais estáveis dentro do arquivo.
    """
    mapa = {}
    proximo = 10001

    # Primeiro os grupos realmente materializados.
    for g in resultado.grupos:
        chave = str(g.grupo_id)
        if chave not in mapa:
            mapa[chave] = proximo
            proximo += 1

    return mapa


def _grupos_numericos(grupos, mapa):
    saida = []
    vistos = set()

    for gid in grupos or []:
        codigo = mapa.get(str(gid))
        if codigo is None or codigo in vistos:
            continue
        vistos.add(codigo)
        saida.append(str(codigo))

    return ",".join(saida)


def gerar_xlsx(template_bytes, resultado):
    """
    Gera XLSX sobre o TEMPLATE OFICIAL.

    Regra crítica para importação:
    - Linha 1 e linha 2 permanecem intactas.
    - Dados começam exatamente na linha 3.
    - Códigos de ITEM e GRUPO são sempre números inteiros.
    - IDs externos (UUID/ObjectId) nunca são enviados ao importador.
    """
    bio = BytesIO(template_bytes)
    wb = load_workbook(bio)

    required = [
        "Item Regular",
        "Grupo de itens adicionais",
        "Item Pesado",
        "Pizza",
    ]
    for s in required:
        if s not in wb.sheetnames:
            raise ValueError(f'Template sem a aba obrigatória "{s}".')

    wi = wb["Item Regular"]
    wg = wb["Grupo de itens adicionais"]
    wp = wb["Pizza"]

    _clear_data(wi, ITEM_HEADERS)
    _clear_data(wg, GRUPO_HEADERS)
    _clear_data(wp, PIZZA_HEADERS)

    mapa_grupos = _montar_mapa_grupos(resultado)

    # Código de item é deliberadamente sequencial e numérico.
    # Isso reproduz o formato dos XLSX de Anota AI que já eram aceitos.
    codigo_item = 1

    for idx, p in enumerate(resultado.itens, start=3):
        if idx > 3:
            _copy_row_style(wi, 3, idx, ITEM_HEADERS)

        vals = [
            codigo_item,
            _grupos_numericos(p.grupos, mapa_grupos),
            p.nome,
            p.descricao,
            p.categoria,
            _imagem_para_importacao(p.imagem),
            "",
            float(p.preco or 0),
            1, 1,
            0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
        ]
        for c, v in enumerate(vals, start=1):
            wi.cell(idx, c).value = v

        codigo_item += 1

    # Cada linha de opção usa o mesmo código NUMÉRICO do seu grupo.
    for idx, g in enumerate(resultado.grupos, start=3):
        if idx > 3:
            _copy_row_style(wg, 3, idx, GRUPO_HEADERS)

        grupo_codigo = mapa_grupos[str(g.grupo_id)]

        vals = [
            grupo_codigo,
            int(g.tipo or 1),
            g.grupo_nome,
            g.nome,
            _imagem_para_importacao(g.imagem),
            float(g.preco or 0),
            int(g.minimo or 0),
            int(g.maximo or 0),
            int(g.repetir or 0),
            int(g.metodo_preco or 1),
        ]
        for c, v in enumerate(vals, start=1):
            wg.cell(idx, c).value = v

    for idx, p in enumerate(resultado.pizzas, start=3):
        if idx > 3:
            _copy_row_style(wp, 3, idx, PIZZA_HEADERS)

        vals = [
            codigo_item,
            _grupos_numericos(p.grupos, mapa_grupos),
            p.nome,
            p.descricao,
            p.categoria,
            _imagem_para_importacao(p.imagem),
            int(p.metodo_preco_pizza or 0),
            float(p.preco or 0),
            1, 1,
            0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
        ]
        for c, v in enumerate(vals, start=1):
            wp.cell(idx, c).value = v

        codigo_item += 1

    # Remove hyperlinks explícitos de todas as abas.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    cell.hyperlink = None

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
