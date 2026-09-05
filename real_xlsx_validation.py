"""Validação controlada de XLSX com dados reais do Leitor Universal V2.

Este script não altera o app nem publica arquivos de cardápio. Para cada plataforma
prioritária, ele reutiliza o mesmo caminho de leitura (parser oficial quando possível,
fallback universal quando necessário), valida o Resultado e gera um XLSX em memória
sobre um template mínimo estrutural. O XLSX é reaberto para confirmar integridade e
é inspecionado contra fórmulas e hyperlinks.
"""
from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook, load_workbook
import requests

from brendi_result_enrichment import extrair_nuxt_data_html, enriquecer_resultado_brendi_nuxt
from preview_runner import gerar_previa_universal
from universal_integration import converter_previa_para_resultado
from universal_router import detectar_url, ler_url_universal
from validator import validar
from xlsx_writer import gerar_xlsx


CASOS_XLSX = [
    ("Anota AI", "https://app.anota.ai/m/xPELP5xiw"),
    ("RapidFood", "https://rapidfood.com.br/panelamineira"),
    ("byFood", "https://pointdogosasco.byfood.com.br"),
    ("InstaDelivery", "https://instadelivery.com.br/acaidorafa1"),
    ("Brendi", "https://pedido.brendi.com.br/pizzaria-tortelli/"),
    ("Ola Click", "https://la-petite-5.ola.click/products"),
    ("Saipos", "https://xisda15.saipos.com/home"),
    ("Cardapio Web", "https://app.cardapioweb.com/shakepoint_westplaza"),
]

# Guardas estruturais só entram quando há evidência pública inequívoca na própria
# fonte real usada pela bateria. A loja Brendi de controle expõe pizzas que exigem
# escolha de sabores; portanto, zero grupos/opções é perda estrutural e não pode
# ser contabilizado como XLSX saudável apenas porque o arquivo abre corretamente.
MINIMOS_ESTRUTURAIS = {
    "Brendi": {"opcoes": 1, "produtos_com_grupo": 1},
}


def _template_minimo() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Regular"
    for nome in ("Grupo de itens adicionais", "Item Pesado", "Pizza"):
        wb.create_sheet(nome)
    for sheet in wb.worksheets:
        sheet.cell(1, 1).value = "linha preservada 1"
        sheet.cell(2, 1).value = "cabecalho preservado"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _resultado_por_url(url: str):
    det = detectar_url(url)
    erro_oficial = None

    if det.estrategia != "diagnostico":
        try:
            resultado, _ = ler_url_universal(det.url_normalizada, usar_playwright=True)
            if resultado is not None and (resultado.itens or resultado.pizzas):
                return resultado, "parser-oficial", erro_oficial
        except Exception as exc:
            erro_oficial = f"{type(exc).__name__}: {exc}"

    previa = gerar_previa_universal(det.url_normalizada)
    if not (previa.produtos or []):
        return None, "fallback-generico", erro_oficial
    resultado = converter_previa_para_resultado(previa, exigir_aprovacao=True)
    return resultado, "fallback-generico", erro_oficial


def _enriquecer_brendi_real(resultado, url: str) -> Dict[str, Any]:
    """Exercita a correção Brendi em ambiente de validação, sem alterar o app.

    A leitura é fail-closed: qualquer mudança na página pública ou no Nuxt faz a
    bateria falhar, em vez de inventar sabores, vínculos ou preços.
    """
    resposta = requests.get(
        url,
        timeout=25,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/131.0 Safari/537.36"
            )
        },
    )
    resposta.raise_for_status()
    nuxt_data = extrair_nuxt_data_html(resposta.text)
    _, auditoria = enriquecer_resultado_brendi_nuxt(resultado, nuxt_data)
    if int(auditoria.get("produtos_vinculados") or 0) < 1:
        raise ValueError("Brendi: enriquecimento real não vinculou nenhum produto de pizza.")
    if int(auditoria.get("opcoes_materializadas") or 0) < 1:
        raise ValueError("Brendi: enriquecimento real não materializou sabores.")
    return auditoria


def _inspecionar_xlsx(xlsx: bytes) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(xlsx), data_only=False)
    formulas = 0
    hyperlinks = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink is not None:
                    hyperlinks += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
    return {
        "bytes": len(xlsx),
        "abas": wb.sheetnames,
        "formulas": formulas,
        "hyperlinks": hyperlinks,
    }


def _regressao_estrutural(label: str, item: Dict[str, Any]) -> str | None:
    minimo = MINIMOS_ESTRUTURAIS.get(label)
    if not minimo:
        return None

    faltas = []
    for campo, esperado in minimo.items():
        atual = int(item.get(campo) or 0)
        if atual < esperado:
            faltas.append(f"{campo}={atual} (mínimo comprovado: {esperado})")
    if not faltas:
        return None
    return "; ".join(faltas)


def validar_caso_xlsx(label: str, url: str) -> Dict[str, Any]:
    item: Dict[str, Any] = {"caso": label, "url": url}
    try:
        resultado, caminho, erro_oficial = _resultado_por_url(url)
        item["caminho_leitura"] = caminho
        if erro_oficial:
            item["erro_parser_oficial"] = erro_oficial
        if resultado is None:
            item.update({"status": "sem-produtos", "xlsx_gerado": False})
            return item

        if label == "Brendi":
            item["enriquecimento_brendi"] = _enriquecer_brendi_real(resultado, url)
            item["caminho_leitura"] = caminho + "+nuxt-pizzas-validacao"

        produtos = list(resultado.itens or []) + list(resultado.pizzas or [])
        item.update({
            "produtos": len(produtos),
            "pizzas": len(resultado.pizzas or []),
            "opcoes": len(resultado.grupos or []),
            "produtos_com_grupo": sum(1 for p in produtos if p.grupos),
            "produtos_com_imagem": sum(1 for p in produtos if p.imagem),
        })

        regressao = _regressao_estrutural(label, item)
        if regressao:
            item.update({
                "status": "regressao-estrutura",
                "xlsx_gerado": False,
                "erro": (
                    "A fonte real comprova estrutura configurável que não foi preservada pelo leitor: "
                    + regressao
                ),
            })
            return item

        erros, avisos = validar(resultado)
        item["avisos_validacao"] = list(avisos or [])[:10]
        if erros:
            item.update({
                "status": "nao-elegivel",
                "xlsx_gerado": False,
                "erros_validacao": list(erros)[:10],
            })
            return item

        xlsx = gerar_xlsx(_template_minimo(), resultado)
        inspecao = _inspecionar_xlsx(xlsx)
        if inspecao["formulas"] or inspecao["hyperlinks"]:
            item.update({
                "status": "erro-xlsx",
                "xlsx_gerado": True,
                "inspecao": inspecao,
                "erro": "XLSX gerado contém fórmula ou hyperlink inesperado.",
            })
            return item

        item.update({
            "status": "xlsx-ok",
            "xlsx_gerado": True,
            "inspecao": inspecao,
        })
        return item
    except Exception as exc:
        item.update({
            "status": "erro-xlsx",
            "xlsx_gerado": False,
            "erro": f"{type(exc).__name__}: {exc}",
        })
        return item


def main() -> int:
    resultados = []
    for label, url in CASOS_XLSX:
        print(f"[XLSX real] {label}", flush=True)
        r = validar_caso_xlsx(label, url)
        resultados.append(r)
        print(
            f"[XLSX real] {label}: {r.get('status')} / "
            f"{r.get('produtos', 0)} produto(s) / {r.get('opcoes', 0)} opção(ões)",
            flush=True,
        )

    payload = {
        "template": "minimo-estrutural-em-memoria",
        "total_casos": len(resultados),
        "xlsx_ok": sum(1 for r in resultados if r.get("status") == "xlsx-ok"),
        "nao_testaveis": sum(1 for r in resultados if r.get("status") == "sem-produtos"),
        "nao_elegiveis": sum(1 for r in resultados if r.get("status") == "nao-elegivel"),
        "regressoes_estrutura": sum(1 for r in resultados if r.get("status") == "regressao-estrutura"),
        "falhas_xlsx": sum(1 for r in resultados if r.get("status") == "erro-xlsx"),
        "resultados": resultados,
    }
    Path("artifacts").mkdir(exist_ok=True)
    Path("artifacts/real_xlsx_report.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps({
        k: payload[k]
        for k in (
            "total_casos", "xlsx_ok", "nao_testaveis", "nao_elegiveis",
            "regressoes_estrutura", "falhas_xlsx"
        )
    }, ensure_ascii=False))
    return 1 if payload["falhas_xlsx"] or payload["regressoes_estrutura"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
